#!/usr/bin/env python3

import argparse
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pdfplumber
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


FINANCIAL_KEYWORDS = [
    # Statements
    "balance sheet",
    "statement of operations",
    "income statement",
    "statement of income",
    "cash flow",
    "cash flows",
    "comprehensive income",
    # Common finance terms
    "revenue",
    "sales",
    "gross profit",
    "gross margin",
    "operating income",
    "operating loss",
    "operating margin",
    "net income",
    "net loss",
    "eps",
    "earnings per share",
    "ebit",
    "ebitda",
    "cogs",
    "cost of goods",
    "operating expenses",
    "r&d",
    "research and development",
    "sg&a",
    "liabilities",
    "assets",
    "equity",
    "stockholders' equity",
    "shareholders' equity",
    "free cash flow",
    "gaap",
    "non-gaap",
]


def cell_has_letters(value: Optional[str]) -> bool:
    if value is None:
        return False
    return bool(re.search(r"[A-Za-z]", str(value)))


def _row_len_maybe_scalar(row: object) -> int:
    """Return number of logical cells in a row.

    Some PDF extraction strategies occasionally return a scalar string for an
    entire row (e.g., "HIGHLIGHTS"). In those cases, we consider the row to have
    a single cell rather than splitting the string into characters.
    """
    if isinstance(row, (list, tuple)):
        return len(row)  # already a list of cells
    # scalar (str/None/number) -> treat as single cell
    return 1


def _ensure_row_list(row: object) -> List[Optional[str]]:
    """Convert an extracted row into a list of cells without splitting strings.

    If a row is a scalar string (or any non-list type), wrap it in a list so we
    append the whole sentence to a single cell in Excel.
    """
    if isinstance(row, (list, tuple)):
        return list(row)
    return [row]  # single cell containing the full value


def normalize_rows(rows: List[List[Optional[str]]]) -> List[List[Optional[str]]]:
    """Pad rows to have equal length and keep sentences intact.

    - Ensures scalar rows are treated as a single cell (prevents character-
      by-character splitting in Excel).
    - Applies basic trimming while preserving None values.
    """
    if not rows:
        return rows

    max_len = max(_row_len_maybe_scalar(r) for r in rows)
    out: List[List[Optional[str]]] = []
    for r in rows:
        cells = _ensure_row_list(r)
        padded = list(cells) + [None] * (max_len - len(cells))
        cleaned = [str(c).strip() if c is not None else None for c in padded]
        out.append(cleaned)
    return out


def detect_header_row(rows: List[List[Optional[str]]]) -> bool:
    if len(rows) < 2:
        return False
    first_row = rows[0]
    second_row = rows[1]
    first_row_letter_ratio = sum(cell_has_letters(c) for c in first_row) / max(1, len(first_row))
    second_row_letter_ratio = sum(cell_has_letters(c) for c in second_row) / max(1, len(second_row))
    return first_row_letter_ratio >= second_row_letter_ratio


def extract_tables_from_page(page: pdfplumber.page.Page) -> List[List[List[Optional[str]]]]:
    strategies: List[Dict[str, str]] = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
        {"vertical_strategy": "lines_strict", "horizontal_strategy": "lines_strict"},
        {"vertical_strategy": "text", "horizontal_strategy": "text"},
        {"vertical_strategy": "lines", "horizontal_strategy": "text"},
        {"vertical_strategy": "text", "horizontal_strategy": "lines"},
    ]

    try:
        found = page.find_tables(table_settings=strategies[0])
        tables = [t.extract() for t in found if t is not None]
        if tables:
            return tables
    except Exception:
        pass

    for settings in strategies:
        try:
            tables = page.extract_tables(table_settings=settings)
            if tables:
                return tables
        except Exception:
            continue

    return []


def extract_tables_from_pdf(pdf_path: Path, min_columns: int = 2) -> List[Tuple[int, List[List[Optional[str]]]]]:
    results: List[Tuple[int, List[List[Optional[str]]]]] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            raw_tables = extract_tables_from_page(page)
            for raw in raw_tables:
                norm = normalize_rows(raw)
                # Filter very small/degenerate tables
                if not norm:
                    continue
                if len(norm[0]) < min_columns:
                    continue
                results.append((page_idx, norm))
    return results


def is_financial_table_rows(rows: List[List[Optional[str]]]) -> bool:
    if not rows:
        return False
    sample_lines: List[str] = []
    # columns header + first up to 3 rows
    sample_lines.append(" ".join([str(c) for c in rows[0]]))
    for i in range(1, min(4, len(rows))):
        sample_lines.append(" ".join([str(c) for c in rows[i]]))
    text_sample = "\n".join(sample_lines).lower()
    return any(keyword in text_sample for keyword in FINANCIAL_KEYWORDS)


def collect_pdfs(inputs: List[str]) -> List[Path]:
    pdfs: List[Path] = []
    for inp in inputs:
        p = Path(inp)
        if p.is_file() and p.suffix.lower() == ".pdf":
            pdfs.append(p)
        elif p.is_dir():
            pdfs.extend(sorted(p.glob("**/*.pdf")))
        else:
            matched = list(Path().glob(inp))
            pdfs.extend([m for m in matched if m.suffix.lower() == ".pdf"])
    seen = set()
    unique_pdfs: List[Path] = []
    for p in pdfs:
        if p.resolve() not in seen:
            unique_pdfs.append(p)
            seen.add(p.resolve())
    return unique_pdfs


def sanitize_sheet_name(name: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = ''.join('_' if c in invalid else c for c in name)
    return cleaned[:31] if len(cleaned) > 31 else cleaned


def make_unique_sheet_name(base: str, used: set) -> str:
    base = sanitize_sheet_name(base)
    name = base
    counter = 1
    while name in used or not name:
        suffix = f"_{counter}"
        limit = 31 - len(suffix)
        name = (base[:limit] if len(base) > limit else base) + suffix
        counter += 1
    used.add(name)
    return name


def autosize_columns(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            length = len(str(value)) if value is not None else 0
            widths[idx] = max(widths.get(idx, 0), length)
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 60)


def _coerce_numeric(value: Optional[str]):
    """Convert common numeric strings to numbers when possible.

    Handles:
    - Thousands separators and currency symbols ($, €, £)
    - Percent values (kept as numeric, e.g., '7.6%' -> 7.6)
    - K/M/B suffixes (1.2M -> 1_200_000)
    - Parentheses for negatives: (123) -> -123
    Returns the converted number or the original value if not parseable.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None

    # Normalize unicode minus and whitespace
    s_clean = s.replace("\u2212", "-")
    # Remove currency symbols and spaces
    s_clean = re.sub(r"[\s$€£]", "", s_clean)

    # Detect negative in parentheses
    negative = False
    if s_clean.startswith("(") and s_clean.endswith(")"):
        negative = True
        s_clean = s_clean[1:-1]

    # Handle suffix multipliers
    multiplier = 1.0
    if re.search(r"[Kk]$", s_clean):
        multiplier = 1_000.0
        s_clean = s_clean[:-1]
    elif re.search(r"[Mm]$", s_clean):
        multiplier = 1_000_000.0
        s_clean = s_clean[:-1]
    elif re.search(r"[Bb]$", s_clean):
        multiplier = 1_000_000_000.0
        s_clean = s_clean[:-1]

    # Percent
    is_percent = False
    if s_clean.endswith("%"):
        is_percent = True
        s_clean = s_clean[:-1]

    # Remove thousands separators
    s_clean = s_clean.replace(",", "")

    # Sometimes there are footnote markers like "1" or "†" at the end
    s_clean = re.sub(r"[^0-9.+\-]", "", s_clean)

    try:
        num = float(s_clean) * multiplier
        if negative:
            num = -num
        # Keep percents as their numeric value (e.g., 7.6)
        return num
    except Exception:
        return value


def write_tables_to_excel(
    extracted: List[Tuple[Path, List[Tuple[int, List[List[Optional[str]]]]]]],
    output_path: Path,
    only_financial: bool,
) -> None:
    """Write all extracted tables into two sheets: reference and financials.

    - Keeps sentences in the same cells (no character splitting)
    - Aggregates ALL non-financial tables into a single "reference" sheet
    - Aggregates ALL financial tables into a single "financials" sheet
    - Adds lightweight separators with source metadata between tables
    """
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_ref = wb.create_sheet(title="reference")
    ws_fin = wb.create_sheet(title="financials")

    def append_table(ws, header: str, rows: List[List[Optional[str]]]):
        # Separator and source header
        if ws.max_row > 1:
            ws.append([None])
        ws.append([header])
        for r in rows:
            # Safety: ensure a string row is treated as a single cell, not split into characters
            r = _ensure_row_list(r)
            ws.append([
                None if (c is None or str(c).strip() == "") else _coerce_numeric(c)
                for c in r
            ])

    for pdf_path, page_tables in extracted:
        for idx, (page_num, rows) in enumerate(page_tables, start=1):
            # Normalize rows first to avoid character splitting
            rows = normalize_rows(rows)

            is_fin = is_financial_table_rows(rows)
            if only_financial and not is_fin:
                continue
            header = f"Source: {pdf_path.name}  page {page_num}  table {idx}"
            if is_fin:
                append_table(ws_fin, header, rows)
            else:
                append_table(ws_ref, header, rows)

    autosize_columns(ws_ref)
    autosize_columns(ws_fin)

    # Save with only the two required sheets
    wb.save(output_path)


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Extract all tables from one or more PDFs into an Excel workbook.",
    )
    parser.add_argument(
        "--input", "-i", nargs="+", required=True,
        help="PDF files, directories, or glob patterns (space-separated)",
    )
    parser.add_argument(
        "--output", "-o", required=False,
        help="Output .xlsx path. Default: <single-pdf>-tables.xlsx or combined-tables.xlsx",
    )
    parser.add_argument(
        "--financial-only", action="store_true",
        help="Keep only likely financial tables (heuristic keyword filter)",
    )
    parser.add_argument(
        "--min-columns", type=int, default=2,
        help="Minimum number of columns to keep a table (default: 2)",
    )

    args = parser.parse_args(list(argv) if argv is not None else None)

    pdfs = collect_pdfs(args.input)
    if not pdfs:
        raise SystemExit("No PDFs found for given inputs.")

    output = (
        Path(args.output)
        if args.output
        else (
            Path(pdfs[0].with_name(f"{pdfs[0].stem}-tables.xlsx"))
            if len(pdfs) == 1
            else Path("combined-tables.xlsx")
        )
    )
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    extracted: List[Tuple[Path, List[Tuple[int, List[List[Optional[str]]]]]]] = []
    for pdf_path in tqdm(pdfs, desc="Extracting tables from PDFs"):
        tables = extract_tables_from_pdf(pdf_path, min_columns=args.min_columns)
        extracted.append((pdf_path, tables))

    write_tables_to_excel(extracted, output, only_financial=args.financial_only)
    print(f"Saved tables to: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
