#!/usr/bin/env python3

import argparse
import base64
import io
import json
import os
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, Union

import pypdfium2 as pdfium
from PIL import Image
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


FINANCIAL_KEYWORDS: List[str] = [
    # Statements
    "balance sheet",
    "statement of operations",
    "income statement",
    "statement of income",
    "cash flow",
    "cash flows",
    "comprehensive income",
    # Common finance terms
    "revenue",
    "sales",
    "gross profit",
    "gross margin",
    "operating income",
    "operating loss",
    "operating margin",
    "net income",
    "net loss",
    "eps",
    "earnings per share",
    "ebit",
    "ebitda",
    "cogs",
    "cost of goods",
    "operating expenses",
    "r&d",
    "research and development",
    "sg&a",
    "liabilities",
    "assets",
    "equity",
    "stockholders' equity",
    "shareholders' equity",
    "free cash flow",
    "gaap",
    "non-gaap",
]


def autosize_columns(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            length = len(str(value)) if value is not None else 0
            widths[idx] = max(widths.get(idx, 0), length)
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 60)


def collect_pdfs(inputs: List[str]) -> List[Path]:
    pdfs: List[Path] = []
    for inp in inputs:
        p = Path(inp)
        if p.is_file() and p.suffix.lower() == ".pdf":
            pdfs.append(p)
        elif p.is_dir():
            pdfs.extend(sorted(p.glob("**/*.pdf")))
        else:
            matched = list(Path().glob(inp))
            pdfs.extend([m for m in matched if m.suffix.lower() == ".pdf"])
    seen = set()
    unique_pdfs: List[Path] = []
    for p in pdfs:
        if p.resolve() not in seen:
            unique_pdfs.append(p)
            seen.add(p.resolve())
    return unique_pdfs


def render_pdf_page_to_image(pdf_path: Path, page_index: int, dpi: int = 200) -> Image.Image:
    pdf = pdfium.PdfDocument(str(pdf_path))
    try:
        page = pdf[page_index]
        try:
            scale = dpi / 72.0
            bitmap = page.render(scale=scale)
            pil_image: Image.Image = bitmap.to_pil()
            return pil_image
        finally:
            page.close()
    finally:
        pdf.close()


def pil_image_to_data_url(img: Image.Image, fmt: str = "PNG") -> str:
    buf = io.BytesIO()
    img.save(buf, format=fmt)
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/{fmt.lower()};base64,{b64}"


def build_messages_for_graph_extraction(image_data_url: str) -> List[Dict[str, object]]:
    # The prompt content mirrors the user's requested format
    content: List[Dict[str, object]] = [
        {
            "type": "text",
            "text": (
                "You are an assistant that find charts, graphs, or diagrams from an image "
                "and summarize their information. There could be multiple diagrams in one image, "
                "so explain each one of them separately. ignore tables."
            ),
        },
        {
            "type": "text",
            "text": (
                'The response must be a JSON in following format {"graphs": [ , , ]} '
                "where , , and  placeholders that describe each graph found in the image. "
                "Do not append or add anything other than the JSON format response."
            ),
        },
        {
            "type": "text",
            "text": (
                'If could not find a graph in the image, return an empty list JSON as follows: '
                '{"graphs": []}. Do not append or add anything other than the JSON format response. '
                'Dont use coding "```" marks or the word json.'
            ),
        },
        {
            "type": "text",
            "text": (
                "Look at the attached image and describe all the graphs inside it in JSON format. "
                "ignore tables and be concise."
            ),
        },
        {"type": "image_url", "image_url": {"url": image_data_url}},
    ]
    return [{"role": "user", "content": content}]


def extract_graphs_from_response_text(text: str) -> List[Union[str, Dict[str, object]]]:
    # Try strict JSON first
    try:
        data = json.loads(text)
        graphs = data.get("graphs", [])
        if isinstance(graphs, list):
            return graphs
    except Exception:
        pass

    # Try to find the first JSON object containing "graphs"
    # This allows recovery if the model added stray text or code fences
    match = re.search(r"\{[\s\S]*?\}\s*\Z", text)
    if match:
        candidate = match.group(0)
        try:
            data = json.loads(candidate)
            graphs = data.get("graphs", [])
            if isinstance(graphs, list):
                return graphs
        except Exception:
            pass

    # Try to strip code fences if present
    stripped = text.strip()
    if stripped.startswith("```") and stripped.endswith("```"):
        inner = stripped.strip("`")
        try:
            data = json.loads(inner)
            graphs = data.get("graphs", [])
            if isinstance(graphs, list):
                return graphs
        except Exception:
            pass

    return []


def graph_text_from_item(item: Union[str, Dict[str, object]]) -> str:
    if isinstance(item, str):
        return item
    if isinstance(item, dict):
        # Concatenate all primitive values for a textual fingerprint
        parts: List[str] = []
        for v in item.values():
            if isinstance(v, (str, int, float)):
                parts.append(str(v))
            elif isinstance(v, list):
                parts.extend(str(x) for x in v if isinstance(x, (str, int, float)))
        return "; ".join(parts)
    return str(item)


def is_financial_text(text: str) -> bool:
    lt = text.lower()
    return any(k in lt for k in FINANCIAL_KEYWORDS)


def write_results_to_excel(
    results: List[Tuple[Path, int, int, str, bool]],
    output_path: Path,
) -> None:
    """Write results into two sheets: 'reference' and 'financials'.

    Each row: [pdf, page, graph_index, description].
    """
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_ref = wb.create_sheet(title="reference")
    ws_fin = wb.create_sheet(title="financials")

    headers = ["pdf", "page", "graph_index", "description"]
    # Ensure headers are appended as a row, not split into characters
    ws_ref.append(list(headers))
    ws_fin.append(list(headers))

    for pdf_path, page_num, graph_idx, desc, is_fin in results:
        # Safety: coerce any accidental scalar row into a list so strings aren't split
        row = [pdf_path.name, page_num, graph_idx, desc]
        if is_fin:
            ws_fin.append(row)
        else:
            ws_ref.append(row)

    autosize_columns(ws_ref)
    autosize_columns(ws_fin)
    wb.save(output_path)


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "RAG-style: Analyze PDF pages as images with OpenAI Vision to extract graph summaries "
            "and save them to an Excel workbook with 'reference' and 'financials' sheets."
        )
    )
    parser.add_argument("--input", "-i", nargs="+", required=True, help="PDF files, directories, or glob patterns")
    parser.add_argument("--output", "-o", required=False, help="Output .xlsx path (default based on first input)")
    parser.add_argument("--model", default="gpt-4o-mini", help="OpenAI model (vision-capable)")
    parser.add_argument("--max-tokens", type=int, default=1000, help="Max tokens for completion (default: 1000)")
    parser.add_argument("--api-key", help="OpenAI API key (or set OPENAI_API_KEY env var)")
    parser.add_argument("--dpi", type=int, default=200, help="Page render DPI (default: 200)")

    args = parser.parse_args(list(argv) if argv is not None else None)

    output = None
    pdfs = collect_pdfs(args.input)
    if not pdfs:
        raise SystemExit("No PDFs found for given inputs.")

    if args.output:
        output = Path(args.output)
    else:
        output = (
            Path(pdfs[0].with_name(f"{pdfs[0].stem}-graphs.xlsx"))
            if len(pdfs) == 1
            else Path("combined-graphs.xlsx")
        )
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    api_key = args.api_key or os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise SystemExit("Missing OpenAI API key. Set --api-key or OPENAI_API_KEY env var.")

    # Import here to avoid importing if user only inspects CLI help
    from openai import OpenAI

    client = OpenAI(api_key=api_key)

    results: List[Tuple[Path, int, int, str, bool]] = []

    for pdf_path in pdfs:
        # Get page count via pdfium quickly
        pdf = pdfium.PdfDocument(str(pdf_path))
        try:
            num_pages = len(pdf)
        finally:
            pdf.close()

        for page_idx in tqdm(range(num_pages), desc=f"Analyzing {pdf_path.name}"):
            pil = render_pdf_page_to_image(pdf_path, page_idx, dpi=args.dpi)
            data_url = pil_image_to_data_url(pil)
            messages = build_messages_for_graph_extraction(data_url)

            try:
                response = client.chat.completions.create(
                    model=args.model,
                    messages=messages,
                    temperature=0,
                    max_tokens=args.max_tokens,
                )
                content = response.choices[0].message.content or ""
            except Exception:
                # If one page fails, continue with others
                content = "{\"graphs\": []}"

            graphs = extract_graphs_from_response_text(content)
            if not graphs:
                continue

            for idx, g in enumerate(graphs, start=1):
                desc = graph_text_from_item(g)
                is_fin = is_financial_text(desc)
                results.append((pdf_path, page_idx + 1, idx, desc, is_fin))

    write_results_to_excel(results, output)
    print(f"Saved graph summaries to: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
